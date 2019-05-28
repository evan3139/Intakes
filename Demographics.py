import os
from tkinter import Tk
from tkinter.filedialog import askopenfilename

import sys
import openpyxl
from openpyxl.utils import *
from openpyxl.utils import get_column_letter
import pandas as pandas
import pandas as pd
import xlrd
import xlsxwriter
from docx import Document


def demographics(excel_name, trigger):
    # This reopens the excel file but in the openpyxl library allowing us to alter column lengths
    wb = openpyxl.load_workbook(excel_name, data_only=True)
    worksheet = wb.active

    for col in worksheet.columns:
        column = col[0].value  # Get the Column Name Here
        if column == trigger:
            idx = 0
            for index, cell in enumerate(col):
                idx += 1
                try:
                    if index >= 1 and (col[index].value != col[index + 1].value):
                        # Have to add this idx iteration due to "cols" reading the old column number count
                        # If you dont use a custom index counter then it will break.
                        idx +=1
                        worksheet.insert_rows(idx, 1)
                except:
                    pass

    wb.save(excel_name)


def demographics_number_groups(excel_name, trigger, spots_to_space):
    # This reopens the excel file but in the openpyxl library allowing us to alter column lengths
    wb = openpyxl.load_workbook(excel_name, data_only=True)
    worksheet = wb.active

    for col in worksheet.columns:
        column = col[0].value  # Get the Column Name Here
        if column == trigger:
            idx = 0
            i = 0
            for index, cell in enumerate(col):
                idx += 1
                try:
                    if index >= 1 and (col[index].value > spots_to_space[i]):
                        # Have to add this idx iteration due to "cols" reading the old column number count
                        # If you dont use a custom index counter then it will break.
                        i += 1
                        idx += 1
                        worksheet.insert_rows(idx -1, 1)
                except:
                    pass

    wb.save(excel_name)
