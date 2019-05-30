import openpyxl
from IsNumber import is_number
import math


def get_group_average(excel_name):
    wb = openpyxl.load_workbook(excel_name, data_only=True)
    ws = wb.active

    indexes = []
    trigger = 0

    for index, cell in enumerate(ws['A']):
        if not cell.value:
            indexes.append(index)
    print(indexes)

    sum = [0] * (len(indexes) + 1)
    divisor = [0] * (len(indexes) + 1)
    rows = [None] * (len(indexes) + 1)


    # Auto fill the last row cause it wont read none from the sheet it just ends at the last row.
    rows[len(indexes)] = ws.max_row
    max_col = ws.max_column + 3

    for col in ws.columns:
        # this will start iterating only after it hits the ID section
        if trigger == 1:
            idx = 0
            for index, cell in enumerate(col):
                if not cell.value and cell.value != 0:
                    print(idx)
                    print("FUCK")
                    rows[idx] = index
                    idx += 1
                elif is_number(cell.value):
                    print(cell.value)
                    sum[idx] += cell.value
                    divisor[idx] += 1
                print(cell.value)
        if col[0].value == "ID":
            trigger = 1
        idx = 0


    for index, i in enumerate(sum):
        if divisor[index] == 0:
            average = 0
        else:
            average = math.floor(i / divisor[index])
        ws.cell(row=rows[index], column=max_col - 1).value = "Average"
        ws.cell(row=rows[index], column=max_col).value = average

    wb.save(excel_name)
