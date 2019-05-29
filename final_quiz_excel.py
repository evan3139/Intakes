import openpyxl
import math
from IsNumber import *

from GLOBALS import *


def combined_average(excel_totals):

    wb = openpyxl.load_workbook(excel_totals, data_only=True)
    ws = wb.active

    skip_header = 0


    for row in ws.rows:
        index = 0
        sum = 0
        divisor =0
        if skip_header > 0:
            for cell in row:
                if index > 0:
                    if cell.column == ws.max_column and divisor != 0:
                        cell.value = math.floor(sum/divisor)
                    elif cell.value == None or cell.value == "N/A":
                        continue
                    elif is_number(cell.value):
                        sum += cell.value
                        divisor += 1
                index += 1

        skip_header += 1

    wb.save(excel_totals)




def quiz_totals(excel_data, excel_totals, facility, group):
    read_wb = openpyxl.load_workbook(excel_data, data_only=True)
    read_ws = read_wb.active
    wb = openpyxl.load_workbook(excel_totals, data_only=True)
    ws = wb.active

    order = []
    case_insensitive_groups = []

    for groupings in GROUPING_NAMES:
        case_insensitive_groups.append(groupings.lower())

    # Get the index of each group (White, Black , Ages, Education years Etc)
    for cell in ws['A']:
        order.append(cell.value)


    groupings = []
    trigger = 0

    for x, col in enumerate(read_ws.columns):
        if col[0].value == group:
            for i, cell in enumerate(col):
                if not cell.value:
                    # i is acting as a decrement here since getting a cell starts at 1
                    groupings.append(read_ws.cell(row=i, column=x + 1).value)
                if i == len(col) - 1:
                    groupings.append(cell.value)
            break

    sum = [0] * (len(groupings) + 1)
    divisor = [0] * (len(groupings) + 1)
    rows = [None] * (len(groupings) + 1)
    averages = []

    for col in read_ws.columns:

        # This saves it from crashing due to the empty lines for the average after
        if col[0].value == None:
            break
        # this will start iterating only after it hits the ID section
        if trigger == 1:
            idx = 0
            for index, cell in enumerate(col):
                if not cell.value:
                    rows[idx] = index
                    idx += 1
                elif is_number(cell.value):
                    sum[idx] += cell.value
                    divisor[idx] += 1
        if col[0].value == "ID":
            trigger = 1

    for index, i in enumerate(sum):
        if divisor[index] == 0:
            averages.append(0)
        else:
            averages.append(math.floor(i / divisor[index]))


    for index,col in enumerate(ws.columns):
        if col[0].value.lower() == facility.lower():
            # Fill all of it with N/A so ones we dont fill go to N/A
            for i,cell in enumerate(col):
                if i == order.index(None) or ws.cell(row=i + 1, column= 1).value in GROUPING_NAMES:
                    continue
                elif not cell.value:
                    cell.value = "N/A"
            for i,x in enumerate(groupings):
                if isinstance(x,int):
                    if x >= 18:
                        if x <= 28:
                            ws.cell(row=order.index("Ages") + 2, column=index + 1).value = averages[i]
                        elif x <=39:
                            ws.cell(row=order.index("Ages") + 3, column=index + 1).value = averages[i]
                        elif x <= 54:
                            ws.cell(row=order.index("Ages") + 4, column=index + 1).value = averages[i]
                        elif x >= 55:
                            ws.cell(row=order.index("Ages") + 5, column=index + 1).value = averages[i]
                    elif x <= 17:
                        continue
                        #THIS WILL BE EDUCATION MAY HAVE TO DO TUPLES SOON IF WE DO CHILDREN AND SHIT BUT FUCK THAT RN
                else:
                    ws.cell(row=order.index(x) + 1, column=index + 1).value = averages[i]

    wb.save(excel_totals)



def format_totals(excel_totals, header):
    wb = openpyxl.load_workbook(excel_totals, data_only=True)
    ws = wb.active

    ws.cell(row=ws.max_row, column=ws.max_column).value = "Ages"

    for group in header:
        ws.cell(row=ws.max_row, column=ws.max_column + 1).value = group

    # skip a column for neatness and put all combined
    ws.cell(row=ws.max_row, column=ws.max_column + 1).value = "All"

    for i, age in enumerate(AGES):
        if i == len(AGES) - 1:
            ws.cell(row=ws.max_row + 1, column=1).value = (str(age) + "+")
        else:
            ws.cell(row=ws.max_row + 1, column=1).value = (str(age) + "-" + str((AGES[i + 1] - 1)))

    group_name = 0

    for group in GROUPINGS:
        # Create the Header
        ws.cell(row=ws.max_row + 2, column=1).value = GROUPING_NAMES[group_name]
        group_name += 1

        for segment in group:
            ws.cell(row=ws.max_row + 1, column=1).value = segment

    wb.save(excel_totals)
